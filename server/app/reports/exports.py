"""Report exports (S3.1, ticket #23): xlsx + pdf.

Both builders consume the SAME generic grid spec the printable page uses
(`views.ViewSpec`), so a report renders identically on screen, paper and
file. Money cells stay exact-decimal STRINGS in the workbook — never
floats (plan/00 money rule); the sheet is RTL for Arabic-first reading.
The PDF embeds the bundled OFL IBM Plex Sans Arabic (P05 fallback face)
with HarfBuzz shaping so Arabic prints shaped on any machine — the
server-side twin of the browser print path (plan/09 P06).
"""
from __future__ import annotations

import io
import re
from pathlib import Path

from fpdf import FPDF
from fpdf.fonts import FontFace
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

FONTS_DIR = Path(__file__).parent / "fonts"
FONT_REGULAR = str(FONTS_DIR / "IBMPlexSansArabic-Regular.ttf")
FONT_BOLD = str(FONTS_DIR / "IBMPlexSansArabic-Bold.ttf")
FONT_FAMILY = "IBMPlexSansArabic"

# page sizes in mm (fpdf2 format names) — mirrors template.py @page sizes
_PAPER_FORMAT = {"A4": "A4", "A5": "A5"}


def _write_text(ws, row: int, col: int, value: str, font: Font | None = None) -> None:
    """Write a cell as TEXT, never a formula (CWE-1236).

    openpyxl stores any string starting with '=' as a live formula — a
    drug/cashier name like '=WEBSERVICE(...)' planted by a low-priv user
    would execute when an admin opens the export. Forcing data_type='s'
    keeps the text verbatim and inert.
    """
    cell = ws.cell(row=row, column=col, value=value)
    if cell.data_type == "f":
        cell.data_type = "s"
    if font is not None:
        cell.font = font


def build_xlsx(
    *,
    title_ar: str,
    title_en: str,
    meta: list[tuple[str, str]],
    columns: list[str],
    rows: list[list],
    foot: list | None = None,
    note: str | None = None,
) -> bytes:
    """Render one report grid as an .xlsx workbook (RTL, exact strings)."""
    wb = Workbook()
    ws = wb.active
    # Excel sheet names: ≤31 chars, none of []:*?/\
    safe_title = re.sub(r"[\[\]:*?/\\]", " ", title_en)[:31].strip() or "Report"
    ws.title = safe_title
    ws.sheet_view.rightToLeft = True

    bold = Font(bold=True)
    row_idx = 1
    _write_text(ws, row_idx, 1, "فارما تاج — PharmaTag", bold)
    row_idx += 1
    _write_text(ws, row_idx, 1, title_ar, bold)
    row_idx += 1
    _write_text(ws, row_idx, 1, title_en)
    row_idx += 2
    for label, value in meta:
        _write_text(ws, row_idx, 1, str(label))
        _write_text(ws, row_idx, 2, str(value))
        row_idx += 1
    row_idx += 1

    header_row = row_idx
    for col_idx, name in enumerate(columns, start=1):
        _write_text(ws, header_row, col_idx, str(name), bold)
    row_idx += 1
    for row in rows:
        for col_idx, value in enumerate(row, start=1):
            # exact-decimal rule: every cell is written as text, never float
            _write_text(ws, row_idx, col_idx, str(value))
        row_idx += 1
    if foot:
        for col_idx, value in enumerate(foot, start=1):
            _write_text(ws, row_idx, col_idx, str(value), bold)
        row_idx += 1
    if note:
        row_idx += 1
        _write_text(ws, row_idx, 1, str(note))

    # readable column widths: widest cell per column, capped
    for col_idx in range(1, len(columns) + 1):
        letter = get_column_letter(col_idx)
        width = max(
            (
                len(str(ws.cell(row=r, column=col_idx).value or ""))
                for r in range(1, row_idx + 1)
            ),
            default=len(str(columns[col_idx - 1])),
        )
        ws.column_dimensions[letter].width = min(max(width + 2, 10), 60)

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def build_pdf(
    *,
    title_ar: str,
    title_en: str,
    meta: list[tuple[str, str]],
    columns: list[str],
    rows: list[list],
    foot: list | None = None,
    note: str | None = None,
    paper: str = "A4",
) -> bytes:
    """Render one report grid as a shaped-Arabic PDF page (A4/A5)."""
    pdf = FPDF(orientation="P", unit="mm", format=_PAPER_FORMAT.get(paper, "A4"))
    pdf.add_font(FONT_FAMILY, style="", fname=FONT_REGULAR)
    pdf.add_font(FONT_FAMILY, style="B", fname=FONT_BOLD)
    pdf.set_text_shaping(True)
    pdf.set_auto_page_break(auto=True, margin=12)
    pdf.add_page()

    # header — brand accent line, then the bilingual title (black on white)
    pdf.set_font(FONT_FAMILY, style="B", size=14)
    pdf.set_text_color(0, 0, 0)
    pdf.cell(0, 8, "فارما تاج — PharmaTag", align="C", new_x="LMARGIN", new_y="NEXT")
    pdf.set_draw_color(0, 0, 0)
    pdf.set_line_width(0.5)
    pdf.line(pdf.l_margin, pdf.get_y(), pdf.w - pdf.r_margin, pdf.get_y())
    pdf.ln(2)
    pdf.set_font(FONT_FAMILY, style="B", size=12)
    pdf.cell(0, 7, title_ar, align="C", new_x="LMARGIN", new_y="NEXT")
    pdf.set_font(FONT_FAMILY, style="", size=9)
    pdf.cell(0, 5, title_en, align="C", new_x="LMARGIN", new_y="NEXT")
    pdf.ln(3)

    # meta block: label right, value left (RTL reading)
    pdf.set_font(FONT_FAMILY, style="", size=10)
    for label, value in meta:
        y = pdf.get_y()
        pdf.cell(0, 5.5, f"{value}", align="L")
        pdf.set_xy(pdf.l_margin, y)
        pdf.cell(0, 5.5, f"{label}", align="R", new_x="LMARGIN", new_y="NEXT")
    pdf.ln(2)

    # the grid — columns reversed so the FIRST logical column is RIGHTMOST
    rev_columns = [str(c) for c in reversed(columns)]
    body = [[str(v) for v in reversed(row)] for row in rows]
    if foot:
        body.append([str(v) for v in reversed(foot)])

    widths = _pdf_column_widths(pdf, rev_columns, body)
    head_style = FontFace(family=FONT_FAMILY, emphasis="BOLD", size_pt=9)
    foot_style = FontFace(family=FONT_FAMILY, emphasis="BOLD", size_pt=9)
    with pdf.table(
        col_widths=widths,
        first_row_as_headings=True,
        repeat_headings=1,
        text_align="RIGHT",
        line_height=5.5,
        padding=1.2,
        headings_style=head_style,
    ) as table:
        head = table.row()
        for name in rev_columns:
            head.cell(name)
        for i, row in enumerate(body):
            is_foot = bool(foot) and i == len(body) - 1
            trow = table.row()
            for value in row:
                trow.cell(value, style=foot_style if is_foot else None)

    if note:
        pdf.ln(2)
        pdf.set_font(FONT_FAMILY, style="", size=8)
        pdf.multi_cell(0, 4.5, note, align="R")

    return bytes(pdf.output())


def _pdf_column_widths(pdf: FPDF, columns: list[str], rows: list[list[str]]) -> tuple:
    """Proportional widths from longest content per column, bounded."""
    printable = pdf.w - pdf.l_margin - pdf.r_margin
    weights = []
    for idx in range(len(columns)):
        cells = [columns[idx]] + [r[idx] for r in rows]
        longest = max((len(c) for c in cells), default=1)
        weights.append(max(min(longest, 40), 3))
    total = sum(weights)
    return tuple(round(printable * w / total, 2) for w in weights)
